from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from email.mime.text import MIMEText
from google.auth.transport.requests import Request
import base64
import os
import openpyxl
import re

def is_valid_email(email):
    regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if re.match(regex, email):
        return True
    else:
        return False

def read_emails_from_excel_file(filename):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        valid_emails = []
        for row in range(1, sheet.max_row + 1):  
            email = sheet.cell(row=row, column=1).value 
            if email and is_valid_email(email): 
                valid_emails.append(email)
        return valid_emails

    except FileNotFoundError:
        print(f"The file {filename} was not found.")
        return []

    except Exception as e:
        print(f"An Error Occurred: {e}")
        return []

def create_message(sender, to, subject, message_text):
    message = MIMEText(message_text)
    message['to'] = to
    message['from'] = sender
    message['subject'] = subject
    raw = base64.urlsafe_b64encode(message.as_bytes()).decode()
    return {'raw': raw}

def send_message(service, user_id, message):
    try:
        message = (service.users().messages().send(userId=user_id, body=message)
                   .execute())
        print(f'Message Sent Successfully Id: {message["id"]}')
        return message
    except HttpError as error:
        print(f'An Error Occurred: {error}')
        return None

def service_build(credentials):
    return build('gmail', 'v1', credentials=credentials)

def get_credentials():
    SCOPES = ['https://www.googleapis.com/auth/gmail.send']
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
            # Save the credentials for the next run
            with open('token.json', 'w') as token:
                token.write(creds.to_json())
    return creds

def main():
    creds = get_credentials()
    try:
        service = service_build(creds)
        sender = "sender-email-address-here"
        subject = "Hello from Python"
        message_text = "Hye This is a test email"
        emails_list = read_emails_from_excel_file('emails.xlsx')
        print(emails_list)
        for to in emails_list: 
            message = create_message(sender, to, subject, message_text)
            send_message(service, "me", message)
    except HttpError as error:
        print(f'An Error Occurred: {error}')

if __name__ == '__main__':
    main()
